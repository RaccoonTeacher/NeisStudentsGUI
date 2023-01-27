#!/usr/bin/env python
# coding: utf-8

# In[6]:


import tkinter as tk #GUI 생성
import tkinter.font #Font 설정
import random #Random, 즉 무작위 활용
import openpyxl #excel 불러오기, 수정, 저장
from datetime import datetime #특정 시점 날짜, 요일 활용

#기본 설정
    #요일 설정. 최종 파일 "history.xlsx"에 해당 요일 입력
days = ["월", "화", "수", "목", "금", "토", "일"]

    #불러오기_학생 이름
wb = openpyxl.load_workbook("studentsNames.xlsx") #학생이름이 담겨있는 excel 파일 열기
ws = wb.active #excel 파일의 첫번째 sheet 선택
names = [] #학생 이름을 담을 list
for row in range(2, ws.max_row+1): #excel의 첫번째 열이 학생 이름. 이름을 names 리스트에 넣기
    names.append((ws.cell(row = row, column = 1).value))
scores = [] #학생 점수를 담을 list. 임시저장을 위해 활용
for row in range(2, ws.max_row+1): #excel의 두번째 열이 발표횟수. 발표횟수를 scores 리스트에 넣기
    scores.append((ws.cell(row = row, column = 2).value))
dict_students = dict(zip(names, scores)) #list를 합해 dictionary 생성. {학생이름1:발표횟수1, 학생이름2:발표횟수2...}

    #불러오기_쿠폰
coupons = [] #쿠폰 종류를 담을 list
file_coupon = open("Coupon.txt", "r", encoding = "UTF-8") #메모장에 사전 입력된 쿠폰 불러와 coupons에 넣기
for coupon in file_coupon.readlines(): 
    coupons.append(coupon.strip())
    
#활용한 함수들
def reset(): #학생 점수를 0점으로 재설정 후 모든 Button에 반영
    for name in list(dict_students.keys()): #dict_students의 학생 발표횟수를 모두 0으로
        dict_students[name] = 0
    for index, btn in enumerate(button): #모든 Button의 숫자 변경
        btn.config(text= names[index] + str(dict_students[names[index]]),
                   font=font_name)
        
def drawing_lots(): #학생 1명 추출. 발표 횟수가 많을 수록 뽑힐 확률이 올라감
    list_students = [] #학생 발표 횟수에 따른 학생 이름 list 생성
    for name, count in dict_students.items(): #학생 발표 횟수에 따라 학생 이름을 list에 넣기. "정인재"가 2번 발표했으면 []"정인재", "정인재"]
        list_students.append([name] * count)
    list_students = sum(list_students, []) #모든 분리된 list를 하나의 list로
    global lucky_student #오늘 뽑힌 단 1명의 학생 이름이 담길 변수 lucky_student
    lucky_student= list_students[random.randint(0, len(list_students)-1)] #랜덤으로 학생 1명 뽑기
    result = ("😁", lucky_student, coupons[random.randint(0, len(coupons)-1)], "!") #학생이름, 쿠폰종류가 담긴 출력문
    lbl_result.config(text=result) #GUI에 표시

def click_button(index): #버튼을 클릭할 때 마다 Button의 학생이름 옆 숫자 1씩 증가
    dict_students[names[index]] += 1
    button[index].config(text=list(dict_students.keys())[index] + str(dict_students[names[index]]))
    
def temporary_save(): #학생 점수 임시 저장. 본 프로그램을 중간 중간에 종료할 수 있도록
    for index, value in enumerate(list(dict_students.items())):
        ws.cell(row = index+2, column = 2, value = value[1])
    wb.save("studentsNames.xlsx")
    window.destroy() #본 프로그램 종료
    
def final_save(): #오늘 날짜, 요일, 학생별 발표 총 횟수, 뽑기 당첨 여부를 excel에 저장
    file_excel = openpyxl.load_workbook("history.xlsx") #저장할 excel 파일 불러오기
    sheet = file_excel.active #excel 파일의 sheet 설정(첫번째 sheet)
    date = datetime.now().strftime("%Y.%m.%d") #오늘 날짜 형태를 2023.01.23.과 같이 출력
    day = days[datetime.now().weekday()] #오늘 요일을 "월"~"일"로 설정
    for name, num in dict_students.items(): #뽑기 당첨된 학생은 뽑기당첨여부가 1로 표시, 그 외에는 0으로 표시. 뽑기의 공정성을 보여주기 위함
        if name == lucky_student:
            sheet.append([date, day, name, num, 1])
        else:
            sheet.append([date, day, name, num, 0])
    file_excel.save("history.xlsx") #수정된 내용을 excel 파일에 저장
    window.destroy() #본 프로그램 종료

#GUI 실행
window = tk.Tk()
font_name=tkinter.font.Font(family="맑은 고딕", size=12) #학생이름 font 설정
font_click=tkinter.font.Font(family="맑은 고딕", size=15, slant="italic", weight="bold") #click 버튼 font 설정
font_result=tkinter.font.Font(family="맑은 고딕", size=15, weight = "bold") #뽑기 결과 출력 font 설정
window.title("발표 횟수를 높이면 당첨 확률이 올라간다!") #본 프로그램 제목
window.geometry("500x750+0+0") #최초 GUI 창 크기
window.resizable(True, True) #창 크기 조절 가능하도록 설정

#학생 이름 버튼 생성
button = [] #모든 버튼 Class가 담길 리스트
label = [] #모든 Label이 담길 리스트
for i in range(len(dict_students)): #학생이름, 발표횟수가 담긴 버튼 생성(eg. 정인재1)
    button.append(tk.Button(window, text=names[i] + str(dict_students[names[i]]),
                            font=font_name,
                            command=lambda i=i: click_button(i)))
    button[i].grid(row=i // 3, column=i % 3, ipadx = 20, ipady=3)
    
#학생 이름 밑에 버튼
location = 1
    
    #뽑기 버튼
tk.Button(window, text="뽑기", command=drawing_lots,
         font=font_click,
         bg="black",
         fg="white").grid(row = (len(names) // 3)+location, column = 1)

    #리셋 버튼
tk.Button(window, text="리셋", command=reset,
         font=font_click,
         bg="black",
         fg="white").grid(row = (len(names) // 3)+location+1, column = 1)

    #임시저장 후 종료 버튼
tk.Button(window, text="임시저장 후 종료", command=temporary_save,
         font=font_click,
         bg="black",
         fg="white").grid(row = (len(names) // 3)+location+2, column = 1)

    #최종 저장 후 종료 버튼
tk.Button(window, text="저장 후 종료", command=final_save,
         font=font_click,
         bg="black",
         fg="white").grid(row = (len(names) // 3)+location+3, column = 1)

    #결과 GUI로 나타내기 버튼
lbl_result = tk.Label(window, text="",
                      bd=5,
                      font=font_result)
lbl_result.grid(row = (len(names) // 3)+location+4,
                column = 1,
                rowspan=3)

#종료
window.mainloop()

